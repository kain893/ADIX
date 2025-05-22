#!/usr/bin/env python3
import csv
import os
from datetime import datetime, timedelta, timezone

from aiogram import Bot, Dispatcher, types
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State

from config import ADMIN_IDS, MARKETING_GROUP_ID, MARKIROVKA_GROUP_ID
from database import SessionLocal, User, Ad, ChatGroup, AdFeedback, Sale, TopUp, Withdrawal
from database import SupportTicket, SupportMessage, AdComplaint
from utils import post_ad_to_chat, rus_status


class AdminStates(StatesGroup):
    remove_group = State()
    add_chat = State()
    edit_ad_v1 = State()
    edit_ad_v2 = State()
    deactivate_ad = State()
    ban_unban_user = State()
    broadcast = State()
    balance_user_id = State()
    balance_value = State()
    edit_profile_user_id = State()
    edit_profile_field = State()
    edit_profile_value = State()
    reply_support_ticket = State()
    complaint_write_seller = State()
    complaint_ban_user = State()
    waiting_for_chats_file = State()

def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS

def register_admin_handlers(bot: Bot, dp: Dispatcher):
    @dp.message(Command("admin"))
    async def admin_menu(message: types.Message):
        if not is_admin(message.chat.id):
            return await bot.send_message(message.chat.id, "Нет прав для доступа к админ-меню.")
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, keyboard=[
            [
                types.KeyboardButton(text="Управление балансом"),
                types.KeyboardButton(text="Последние заказы")
            ],
            [
                types.KeyboardButton(text="Рассылка"),
                types.KeyboardButton(text="Забанить/Разбанить")
            ],
            [
                types.KeyboardButton(text="Редактировать объявления"),
                types.KeyboardButton(text="Удалить объявление")  # <-- добавили здесь
            ],
            [
                types.KeyboardButton(text="Управление чатами"),
                types.KeyboardButton(text="Управление поддержкой")
            ],
            [
                types.KeyboardButton(text="Редактировать профиль пользователя")
            ],
            [
                types.KeyboardButton(text="Главное меню")
            ]
        ])
        return await bot.send_message(message.chat.id, "Админ-меню:", reply_markup=kb)

    # ------------------------------------------------------------------------
    #            УДАЛИТЬ (ДЕАКТИВИРОВАТЬ) ОБЪЯВЛЕНИЕ
    # ------------------------------------------------------------------------
    @dp.message(lambda m: m.text == "Удалить объявление")
    async def admin_deactivate_ad(message: types.Message, state: FSMContext):
        if not is_admin(message.chat.id):
            return
        await state.set_state(AdminStates.deactivate_ad)
        await bot.send_message(message.chat.id, "Введите ID объявления для деактивации:")

    @dp.message(AdminStates.deactivate_ad)
    async def process_admin_deactivate_ad(message: types.Message, state: FSMContext):
        await state.clear()
        chat_id = message.chat.id
        try:
            ad_id = int(message.text.strip())
        except ValueError:
            return await bot.send_message(chat_id, "❌ Некорректный ID.")

        with SessionLocal() as session:
            ad = session.query(Ad).get(ad_id)
            if not ad:
                return await bot.send_message(chat_id, f"❌ Объявление #{ad_id} не найдено.")
            ad.is_active = False
            session.commit()

        await bot.send_message(chat_id, f"✅ Объявление #{ad_id} деактивировано.")
        try:
            return await bot.send_message(ad.user_id,
                                   f"Ваше объявление #{ad_id} было деактивировано администратором.")
        except:
            return None

    # ------------------------------------------------------------------------
    #      Одобрить / Отклонить продление
    # ------------------------------------------------------------------------
    @dp.callback_query(lambda c: c.data.startswith("approve_ext_") or c.data.startswith("reject_ext_"))
    async def handle_extension_request(call: types.CallbackQuery):
        admin_id = call.from_user.id
        if not is_admin(admin_id):
            return await bot.answer_callback_query(call.id, "Нет прав.", show_alert=True)

        parts = call.data.split("_")  # ['approve','ext','123']
        action, _, ad_id_str = parts
        ad_id = int(ad_id_str)

        with SessionLocal() as session:
            ad = session.query(Ad).get(ad_id)
            if not ad:
                return await bot.answer_callback_query(call.id, "Объявление не найдено.", show_alert=True)

            # снимем кнопки под заявкой
            await bot.edit_message_reply_markup(chat_id=call.message.chat.id, message_id=call.message.message_id, reply_markup=None)

            if action == "approve":
                ad.is_active = True
                ad.created_at = datetime.now(timezone.utc)
                session.commit()

                await bot.send_message(admin_id, f"✅ Продление объявления #{ad_id} одобрено.")
                await bot.send_message(ad.user_id, f"Ваше объявление #{ad_id} продлено на 30 дней и снова активно!")
            else:
                await bot.send_message(admin_id, f"❌ Продление объявления #{ad_id} отклонено.")
                await bot.send_message(ad.user_id, f"К сожалению, продление объявления #{ad_id} отклонено администратором.")

        return await bot.answer_callback_query(call.id)

    # ------------------------------------------------------------------------
    #            УПРАВЛЕНИЕ БАЛАНСОМ
    # ------------------------------------------------------------------------
    @dp.message(lambda m: m.text == "Управление балансом")
    async def admin_balance(message: types.Message, state: FSMContext):
        if not is_admin(message.chat.id):
            return
        await state.set_state(AdminStates.balance_user_id)
        await bot.send_message(message.chat.id, "Введите *ID пользователя*:", parse_mode="Markdown")

    @dp.message(AdminStates.balance_user_id)
    async def process_admin_balance_user(message: types.Message, state: FSMContext):
        try:
            tid = int(message.text)
        except:
            await state.clear()
            return await bot.send_message(message.chat.id, "Некорректный ID.")
        await state.update_data(tid=tid)
        await state.set_state(AdminStates.balance_value)
        return await bot.send_message(message.chat.id, "Введите сумму (или +100 / -50 и т.п.):")

    @dp.message(AdminStates.balance_value)
    async def process_admin_balance_value(message: types.Message, state: FSMContext):
        data = await state.get_data()
        await state.clear()
        target_user_id = data.get("tid")
        val_str = message.text.strip()
        with SessionLocal() as session:
            user = session.query(User).filter_by(id=target_user_id).first()
            if not user:
                return await bot.send_message(message.chat.id, "Пользователь не найден.")
            try:
                if val_str.startswith("+") or val_str.startswith("-"):
                    delta = float(val_str)
                    user.balance = float(user.balance) + delta
                else:
                    new_val = float(val_str)
                    user.balance = new_val
                session.commit()
                return await bot.send_message(message.chat.id, "Баланс изменён.")
            except:
                return await bot.send_message(message.chat.id, "Ошибка при обработке баланса.")

    # ------------------------------------------------------------------------
    #            ПОСЛЕДНИЕ ЗАКАЗЫ
    # ------------------------------------------------------------------------
    @dp.message(lambda m: m.text == "Последние заказы")
    async def admin_orders(message: types.Message):
        if not is_admin(message.chat.id):
            return None
        with SessionLocal() as session:
            sales = session.query(Sale).order_by(Sale.created_at.desc()).limit(10).all()
            if not sales:
                return await bot.send_message(message.chat.id, "Заказов нет.")
            for s in sales:
                st_text = rus_status(s.status)
                info = (
                    f"Sale ID: {s.id} | Ad ID: {s.ad_id}\n"
                    f"Покупатель: {s.buyer_id}, Продавец: {s.seller_id}\n"
                    f"Сумма: {s.amount}, Статус: {st_text}\n"
                    f"Дата: {s.created_at}"
                )
                await bot.send_message(message.chat.id, info)
        return await bot.send_message(message.chat.id, "Конец списка.")

    # ------------------------------------------------------------------------
    #            РАССЫЛКА
    # ------------------------------------------------------------------------
    @dp.message(lambda m: m.text == "Рассылка")
    async def admin_broadcast(message: types.Message, state: FSMContext):
        if not is_admin(message.chat.id):
            return
        await state.set_state(AdminStates.broadcast)
        await bot.send_message(message.chat.id, "Текст рассылки:")

    @dp.message(AdminStates.broadcast)
    async def process_admin_broadcast_text(message: types.Message, state: FSMContext):
        await state.clear()
        txt = message.text.strip()
        with SessionLocal() as session:
            # Рассылку шлём только незаблокированным
            users = session.query(User).filter_by(is_banned=False).all()
            for u in users:
                try:
                    await bot.send_message(u.id, txt)
                except:
                    pass
        await bot.send_message(message.chat.id, "Рассылка завершена.")

    # ------------------------------------------------------------------------
    #            ЗАБАНИТЬ/РАЗБАНИТЬ (из меню)
    # ------------------------------------------------------------------------
    @dp.message(lambda m: m.text == "Забанить/Разбанить")
    async def admin_ban_unban(message: types.Message, state: FSMContext):
        if not is_admin(message.chat.id):
            return
        await state.set_state(AdminStates.ban_unban_user)
        await bot.send_message(message.chat.id, "Введите: `user_id ban` или `user_id unban`")

    @dp.message(AdminStates.ban_unban_user)
    async def process_admin_ban_unban(message: types.Message, state: FSMContext):
        await state.clear()
        parts = message.text.split()
        if len(parts) != 2:
            return await bot.send_message(message.chat.id, "Неверный формат. Нужен: <id> ban|unban")
        try:
            uid = int(parts[0])
            action = parts[1]
        except:
            return await bot.send_message(message.chat.id, "Неверные данные.")
        with SessionLocal() as session:
            user = session.query(User).filter_by(id=uid).first()
            if not user:
                return await bot.send_message(message.chat.id, "Пользователь не найден.")
            if action.lower() == "ban":
                user.is_banned = True
            elif action.lower() == "unban":
                user.is_banned = False
                user.ban_reason = None
                user.ban_until = None
            else:
                return await bot.send_message(message.chat.id, "Неизвестная команда (ожидается ban или unban).")
            session.commit()
        return await bot.send_message(message.chat.id, f"Пользователь {uid} -> {action}.")

    # ------------------------------------------------------------------------
    #            РЕДАКТИРОВАТЬ ОБЪЯВЛЕНИЯ
    # ------------------------------------------------------------------------
    @dp.message(lambda m: m.text == "Редактировать объявления")
    async def admin_edit_ads(message: types.Message, state: FSMContext):
        if not is_admin(message.chat.id):
            return
        await state.set_state(AdminStates.edit_ad_v1)
        await bot.send_message(message.chat.id, "Введите: ID_объявления|Новый текст.\nНапример: `12|Новый текст`")

    @dp.message(AdminStates.edit_ad_v1)
    async def process_admin_edit_ad(message: types.Message, state: FSMContext):
        await state.clear()
        if "|" not in message.text:
            return await bot.send_message(message.chat.id, "Неверный формат. Нужно указать `|` между ID и текстом.")
        ad_id_str, new_text = message.text.split("|", 1)
        try:
            ad_id = int(ad_id_str.strip())
        except:
            return await bot.send_message(message.chat.id, "Неверный ID объявления (не число).")
        new_text = new_text.strip()
        with SessionLocal() as session:
            ad_obj = session.query(Ad).filter_by(id=ad_id).first()
            if not ad_obj:
                return await bot.send_message(message.chat.id, "Объявление не найдено.")
            ad_obj.text = new_text
            session.commit()
        return await bot.send_message(message.chat.id, f"Объявление #{ad_id} обновлено.")

    # ------------------------------------------------------------------------
    #            УПРАВЛЕНИЕ ЧАТАМИ
    # ------------------------------------------------------------------------
    @dp.message(lambda m: m.text == "Управление чатами")
    async def admin_manage_chats(message: types.Message):
        if not is_admin(message.chat.id):
            return
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, keyboard=[
            [
                types.KeyboardButton(text="Добавить чат"),
                types.KeyboardButton(text="Список чатов"),
                types.KeyboardButton(text="Удалить чат")
            ],
            [
                types.KeyboardButton(text="Загрузить чаты (Excel/CSV)")
            ],
            [
                types.KeyboardButton(text="Главное меню")
            ]
        ])
        await bot.send_message(message.chat.id, "Управление чатами:", reply_markup=kb)

    @dp.message(lambda m: m.text == "Добавить чат")
    async def admin_add_chat(message: types.Message, state: FSMContext):
        if not is_admin(message.chat.id):
            return
        await state.set_state(AdminStates.add_chat)
        await bot.send_message(message.chat.id, "Введите: chat_id, название, цена\nНапример: `-10012345, МойЧат, 50`")

    @dp.message(AdminStates.add_chat)
    async def process_admin_add_chat(message: types.Message, state: FSMContext):
        await state.clear()
        parts = message.text.split(",")
        if len(parts) != 3:
            return await bot.send_message(message.chat.id, "Неверный формат. Нужно 3 значения: <chat_id>, <название>, <цена>.")
        try:
            chat_id_val = int(parts[0].strip())
            title = parts[1].strip()
            price = float(parts[2].strip())
        except:
            return await bot.send_message(message.chat.id, "Неверные данные (chat_id или price не числа).")

        if abs(price) > 99999999.99:
            return await bot.send_message(message.chat.id, f"Слишком большая цена ({price}). Чат пропущен.")

        with SessionLocal() as session:
            cg = ChatGroup(chat_id=chat_id_val, title=title, price_1=price, is_active=True)
            session.add(cg)
            session.commit()
        return await bot.send_message(message.chat.id, f"Чат '{title}' добавлен!")

    @dp.message(lambda m: m.text == "Список чатов")
    async def admin_list_chats(message: types.Message):
        if not is_admin(message.chat.id):
            return None
        with SessionLocal() as session:
            chats = session.query(ChatGroup).all()
            if not chats:
                return await bot.send_message(message.chat.id, "Чатов нет в базе.")

        def detect_region(title: str) -> str:
            low = title.lower()
            if "москв" in low and "область" not in low:
                return "Москва"
            elif "область" in low:
                return "Московская область"
            else:
                return "Города РФ"

        grouped = {
            "Москва": [],
            "Московская область": [],
            "Города РФ": []
        }
        for c in chats:
            r = detect_region(c.title)
            grouped[r].append(c)

        result_text = "СПИСОК ЧАТОВ:\n"
        for reg_key in ["Москва", "Московская область", "Города РФ"]:
            arr = grouped[reg_key]
            if not arr:
                continue
            result_text += f"\n=== {reg_key.upper()} ===\n"
            for c in arr:
                line = (f"[ID {c.id}] chat_id={c.chat_id}, Название='{c.title}', "
                        f"Цена={c.price}, Активен={c.is_active}\n")
                result_text += line

        async def send_in_chunks(chat_id_val, text, chunk_size=4000):
            idx = 0
            length = len(text)
            while idx < length:
                await bot.send_message(chat_id_val, text[idx:idx+chunk_size])
                idx += chunk_size

        if not result_text.strip():
            return await bot.send_message(message.chat.id, "Чатов нет в базе.")

        await send_in_chunks(message.chat.id, result_text)
        return await bot.send_message(message.chat.id, "Конец списка.")

    @dp.message(lambda m: m.text == "Удалить чат")
    async def admin_delete_chat(message: types.Message, state: FSMContext):
        if not is_admin(message.chat.id):
            return
        kb = types.InlineKeyboardMarkup(inline_keyboard=[[
            types.InlineKeyboardButton(text="Отмена", callback_data="cancel_chat_deletion"),
        ]])
        await state.set_state(AdminStates.remove_group)
        await bot.send_message(message.chat.id, "Введите ID чата (из БД):", reply_markup=kb)

    @dp.message(AdminStates.remove_group)
    async def process_admin_delete_chat(message: types.Message, state: FSMContext):
        await state.clear()
        try:
            db_id = int(message.text.strip())
        except:
            return await bot.send_message(message.chat.id, "Некорректный ID (не число).")
        with SessionLocal() as session:
            cg = session.query(ChatGroup).filter_by(id=db_id).first()
            if not cg:
                return await bot.send_message(message.chat.id, "Чат не найден.")
            session.delete(cg)
            session.commit()
        return await bot.send_message(message.chat.id, "Чат удалён.")

    @dp.message(lambda m: m.text == "Загрузить чаты (Excel/CSV)")
    async def admin_add_chats_from_excel_csv(message: types.Message, state: FSMContext):
        if not is_admin(message.chat.id):
            return
        await state.set_state(AdminStates.waiting_for_chats_file)
        await bot.send_message(
            message.chat.id,
            "Пришлите файл Excel (XLSX) или CSV с данными о чатах.\n\n"
            "Формат XLSX: (chat_id, title, price)\n"
            "Формат CSV: Название, Кол-во участников, Цена1, Цена2, ... (и т.д.)"
        )

    @dp.message(AdminStates.waiting_for_chats_file)
    async def wait_for_document_file(message: types.Message, state: FSMContext):
        await state.clear()
        if not is_admin(message.chat.id):
            return None

        if not message.document:
            return await bot.send_message(message.chat.id, "Это не файл. Повторите команду.")

        file_info = await bot.get_file(message.document.file_id)
        downloaded = await bot.download_file(file_info.file_path)
        filename = message.document.file_name.lower()
        extension = os.path.splitext(filename)[1]
        file_path = f"temp_chats_{message.chat.id}{extension}"
        with open(file_path, "wb") as f:
            f.write(downloaded)

        if extension == ".xlsx":
            return await import_chats_from_excel(file_path, message.chat.id)
        elif extension == ".csv":
            return await import_chats_from_csv(file_path, message.chat.id)
        else:
            await bot.send_message(message.chat.id, "Неизвестный формат. Нужен XLSX или CSV.")
            os.remove(file_path)
            return None

    async def import_chats_from_excel(file_path: str, admin_chat_id: int):
        """
        Импорт чатов из Excel-файла с тремя листами:
          1-й лист — Москва
          2-й лист — Московская область (МО)
          3-й лист — Города РФ

        Формат каждого листа (начиная с 2-й строки):
          A: название
          B: цена за 1 размещение
          C: цена за 5 размещений
          D: цена за 10 размещений
          E: закреп на 1 день
          F: участники
          G: ID (int или строка "🆔 Chat ID: <chat_id>")
        """
        import openpyxl, os
        from database import SessionLocal, ChatGroup

        rows_added = rows_updated = 0

        def to_float(v):
            try:
                return float(v)
            except:
                return 0.0

        def parse_chat_id(cell):
            if isinstance(cell, (int, float)):
                return int(cell)
            s = str(cell)
            if ':' in s:
                s = s.split(':', 1)[1]
            return int(s.strip())

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            await bot.send_message(admin_chat_id, f"❌ Ошибка чтения XLSX: {e}")
            os.remove(file_path)
            return

        if len(wb.sheetnames) < 3:
            await bot.send_message(admin_chat_id, "❌ В файле должно быть минимум 3 листа.")
            os.remove(file_path)
            return

        sheet_to_region = {
            wb.sheetnames[0]: "moscow",
            wb.sheetnames[1]: "mo",
            wb.sheetnames[2]: "rf",
        }

        with SessionLocal() as session:
            for sheet_name, region_code in sheet_to_region.items():
                ws = wb[sheet_name]
                # перебираем строки, начиная со 2-ой, и сразу читаем 7 колонок
                for row_idx, (title_cell, p1, p5, p10, p_pin, part_cell, id_cell) in \
                        enumerate(ws.iter_rows(min_row=2, max_col=7, values_only=True), start=2):

                    # если нет ID — пропускаем
                    if id_cell is None:
                        continue

                    title = str(title_cell or "").strip()
                    price_1 = to_float(p1)  # из колонки B
                    price_5 = to_float(p5)  # из колонки C
                    price_10 = to_float(p10)  # из колонки D
                    price_pin = to_float(p_pin)  # из колонки E
                    try:
                        participants = int(part_cell or 0)
                    except:
                        participants = 0

                    try:
                        chat_id_val = parse_chat_id(id_cell)
                    except:
                        continue

                    # ищем уже существующую запись
                    cg = session.query(ChatGroup).filter_by(chat_id=chat_id_val).first()
                    if cg:
                        # обновляем все поля
                        cg.title = title
                        cg.region = region_code
                        cg.price_1 = price_1
                        cg.price_5 = price_5
                        cg.price_10 = price_10
                        cg.price_pin = price_pin
                        cg.participants = participants
                        cg.is_active = True
                        rows_updated += 1
                    else:
                        # создаём новую
                        new_cg = ChatGroup(
                            chat_id=chat_id_val,
                            title=title,
                            region=region_code,
                            price_1=price_1,
                            price_5=price_5,
                            price_10=price_10,
                            price_pin=price_pin,
                            participants=participants,
                            is_active=True
                        )
                        session.add(new_cg)
                        rows_added += 1

            session.commit()

        # удаляем временный файл
        try:
            os.remove(file_path)
        except:
            pass

        await bot.send_message(
            admin_chat_id,
            f"📥 Импорт завершён.\n"
            f"➕ Добавлено: {rows_added}\n"
            f"✏️ Обновлено: {rows_updated}"
        )


    # --- импорт CSV ---------------------------------------------------------
    async def import_chats_from_csv(file_path: str,
                                    admin_chat_id: int) -> None:
        """
        Импорт / обновление чатов из CSV.

        Поддерживаемые форматы строк
        1) chat_id, title, price
        2) title, * , price1 [, price2 …]

        • Если chat_id указан → ищем / создаём по chat_id.
        • Если chat_id отсутствует → ищем по title.
            ─ не нашли → создаём чат, выдавая новый tech‑chat_id (-1000, -1001, …)
        """

        rows_added = rows_updated = rows_skipped = 0

        with SessionLocal() as session:

            # вычислим «самый маленький» (по модулю) отрицательный chat_id,
            # чтобы дальше генерировать tech‑id:  min_neg - 1
            min_neg_chat_id = session.query(ChatGroup) \
                .with_entities(ChatGroup.chat_id) \
                .filter(ChatGroup.chat_id < 0) \
                .order_by(ChatGroup.chat_id) \
                .first()
            next_tech_id = (min_neg_chat_id[0] if min_neg_chat_id else 0) - 1

            # --- читаем csv ---------------------------------------------------
            with open(file_path, newline='', encoding='utf-8') as fh:
                reader = csv.reader(fh)
                header = next(reader, None)  # пропустим строку заголовка, если она есть

                for row in reader:
                    if not row or all(not c.strip() for c in row):
                        rows_skipped += 1
                        continue

                    # уберём пустые колонки/пробелы
                    cells = [c.strip() for c in row if c.strip()]

                    # --- Определяем, есть chat_id --------------------------------
                    chat_id_val = None
                    title_val = None
                    price_val = None

                    # если первый столбец – число ⇒ это chat_id
                    first = cells[0].lstrip("‑-")  # знак «‑» & обычный минус
                    if first.isdigit():
                        # Формат 1
                        try:
                            chat_id_val = int(cells[0])
                            title_val = cells[1] if len(cells) > 1 else ""
                            price_cell = cells[2] if len(cells) > 2 else ""
                        except IndexError:
                            rows_skipped += 1
                            continue
                    else:
                        # Формат 2  (chat_id отсутствует)
                        title_val = cells[0]
                        price_cell = cells[-1]  # берём последнюю ячейку

                    # --- Парсим цену --------------------------------------------
                    price_cell = price_cell.replace(" ", "").replace(",", ".")
                    try:
                        price_val = float(price_cell)
                    except ValueError:
                        rows_skipped += 1
                        continue
                    if abs(price_val) > 9.99e7:
                        rows_skipped += 1
                        continue

                    # --- Добавление / обновление --------------------------------
                    if chat_id_val is not None:
                        # поиск/создание по chat_id
                        cg = session.query(ChatGroup).filter_by(chat_id=chat_id_val).first()
                        if cg:
                            cg.title = title_val or cg.title
                            cg.price = price_val
                            rows_updated += 1
                        else:
                            session.add(ChatGroup(chat_id=chat_id_val,
                                                  title=title_val or "Без названия",
                                                  price_1=price_val,
                                                  is_active=True))
                            rows_added += 1
                    else:
                        # chat_id нет → ищем по title
                        cg = session.query(ChatGroup).filter_by(title=title_val).first()
                        if cg:
                            cg.price = price_val
                            rows_updated += 1
                        else:
                            # создаём с техническим chat_id
                            session.add(ChatGroup(chat_id=next_tech_id,
                                                  title=title_val,
                                                  price_1=price_val,
                                                  is_active=True))
                            next_tech_id -= 1
                            rows_added += 1

            session.commit()

        # удаляем файл
        try:
            os.remove(file_path)
        except OSError:
            pass

        # --- отчёт ------------------------------------------------------------
        await bot.send_message(
            admin_chat_id,
            f"✅ Импорт CSV завершён.\n"
            f"➕ Добавлено: {rows_added}\n"
            f"✏️ Обновлено: {rows_updated}\n"
            f"⏭️ Пропущено: {rows_skipped}"
        )

    # -------------------------------------------------------------------------
    # ------------------------------------------------------------------------
    #            МОДЕРАЦИЯ ОБЪЯВЛЕНИЙ (approve/reject/edit/publish)
    # ------------------------------------------------------------------------
    @dp.callback_query(lambda call:
        call.data.startswith("approve_ad_") or
        call.data.startswith("reject_ad_") or
        call.data.startswith("edit_ad_") or
        call.data.startswith("publish_ad_") or
        call.data.startswith("approve_publish_ad_")
    )
    async def handle_moderation_callbacks(call: types.CallbackQuery, state: FSMContext):
        if not is_admin(call.from_user.id):
            return await bot.answer_callback_query(call.id, "Нет прав для модерации.", show_alert=True)

        with SessionLocal() as session:
            data = call.data.split("_", 2)
            action = data[0] + "_" + data[1]
            ad_id_str = data[2] if len(data) > 2 else None

            if not ad_id_str:
                return await bot.answer_callback_query(call.id, "Некорректные данные.", show_alert=True)

            try:
                ad_id = int(ad_id_str)
            except:
                return await bot.answer_callback_query(call.id, "Некорректный ID объявления.", show_alert=True)

            ad_obj = session.query(Ad).filter_by(id=ad_id).first()
            if not ad_obj:
                return await bot.answer_callback_query(call.id, "Объявление не найдено.", show_alert=True)

            user_obj = session.query(User).filter_by(id=ad_obj.user_id).first()

            if action == "approve_ad":
                ad_obj.status = "approved"
                session.commit()
                if user_obj:
                    await bot.send_message(ad_obj.user_id, f"Ваше объявление #{ad_obj.id} теперь «{rus_status('approved')}»!")
                return await bot.answer_callback_query(call.id, "Объявление одобрено.")
            elif action == "reject_ad":
                ad_obj.status = "rejected"
                session.commit()
                if user_obj:
                    await bot.send_message(ad_obj.user_id, f"Ваше объявление #{ad_obj.id} «{rus_status('rejected')}» админом.")
                return await bot.answer_callback_query(call.id, "Объявление отклонено.")
            elif action == "edit_ad":
                await bot.answer_callback_query(call.id, "Введите новый текст объявления в ответ на это сообщение.")
                await state.set_state(AdminStates.edit_ad_v2)
                await state.update_data(ad_id=ad_id)
                return await bot.send_message(
                    call.message.chat.id,
                    f"Редактирование объявления #{ad_id}. Введите новый текст:"
                )
            elif action == "publish_ad":
                if ad_obj.status != "approved":
                    return await bot.answer_callback_query(call.id, "Сначала одобрите объявление (approve_ad).", show_alert=True)
                if ad_obj.ad_type == "format2":
                    target_chat = MARKIROVKA_GROUP_ID
                else:
                    target_chat = MARKETING_GROUP_ID

                await post_ad_to_chat(bot, target_chat, ad_obj, user_obj)
                return await bot.answer_callback_query(call.id, "Объявление опубликовано!")
            elif action == "approve_publish_ad":
                ad_obj.status = "approved"
                session.commit()
                if ad_obj.ad_type == "format2":
                    target_chat = MARKIROVKA_GROUP_ID
                else:
                    target_chat = MARKETING_GROUP_ID

                await post_ad_to_chat(bot, target_chat, ad_obj, user_obj)
                if user_obj:
                    await bot.send_message(ad_obj.user_id, f"Ваше объявление #{ad_obj.id} «{rus_status('approved')}» и опубликовано!")
                return await bot.answer_callback_query(call.id, "Объявление одобрено и опубликовано!")
            else:
                return None

    @dp.message(AdminStates.edit_ad_v2)
    async def process_edit_ad_text(message: types.Message, state: FSMContext):
        data = await state.get_data()
        await state.clear()
        ad_id = data.get("ad_id")
        new_text = message.text.strip()
        with SessionLocal() as session:
            ad_obj = session.query(Ad).filter_by(id=ad_id).first()
            if not ad_obj:
                return await bot.send_message(message.chat.id, "Объявление не найдено при редактировании.")
            ad_obj.text = new_text
            session.commit()
        return await bot.send_message(message.chat.id, f"Объявление #{ad_id} отредактировано.")

    # ------------------------------------------------------------------------
    #            ОДОБРИТЬ/ОТКЛОНИТЬ ПОПОЛНЕНИЕ
    # ------------------------------------------------------------------------
    @dp.callback_query(
        lambda call: call.data.startswith("approve_topup_") or call.data.startswith("reject_topup_")
    )
    async def handle_topup_approval(call: types.CallbackQuery):
        if not is_admin(call.from_user.id):
            return await bot.answer_callback_query(call.id, "Нет прав для модерации.", show_alert=True)

        # извлекаем ID заявки
        topup_id = int(call.data.split("_")[-1])
        with SessionLocal() as session:
            topup_obj = session.query(TopUp).filter_by(id=topup_id, status="pending").first()
            if not topup_obj:
                return await bot.answer_callback_query(call.id, "Заявка не найдена или уже обработана.", show_alert=True)

            # подгружаем пользователя
            user_obj = session.query(User).filter_by(id=topup_obj.user_id).first()
            user_name = f"@{user_obj.username}" if user_obj and user_obj.username else str(user_obj.id)

            pay_sys = getattr(topup_obj, "payment_system", "не указана")
            pay_card = getattr(topup_obj, "card_number", "не указана")

            # убираем кнопки одобрения/отклонения под заявкой
            await bot.edit_message_reply_markup(chat_id=call.message.chat.id, message_id=call.message.message_id, reply_markup=None)

            if call.data.startswith("approve_topup_"):
                # зачисляем средства
                if user_obj:
                    user_obj.balance = float(user_obj.balance) + float(topup_obj.amount)
                topup_obj.status = "approved"
                session.commit()

                await bot.answer_callback_query(call.id, "Пополнение одобрено.")
                await bot.send_message(
                    call.message.chat.id,
                    (
                        f"✅ Пополнение #{topup_id} на сумму {topup_obj.amount} руб. одобрено.\n"
                        f"Пользователь: {user_name}\n"
                        f"Система: {pay_sys}, Карта: {pay_card}\n"
                        f"Новый баланс: {user_obj.balance if user_obj else 'N/A'} руб."
                    )
                )
                # уведомляем пользователя
                if user_obj:
                    await bot.send_message(
                        user_obj.id,
                        (
                            f"Ваше пополнение #{topup_id} на сумму {topup_obj.amount} руб. "
                            f"«{rus_status('approved')}».\n"
                            f"Новый баланс: {user_obj.balance} руб."
                        )
                    )

            else:  # отклонение
                topup_obj.status = "rejected"
                session.commit()

                await bot.answer_callback_query(call.id, "Пополнение отклонено.")
                await bot.send_message(
                    call.message.chat.id,
                    (
                        f"❌ Пополнение #{topup_id} пользователем {user_name} "
                        f"(Система: {pay_sys}, Карта: {pay_card}) «{rus_status('rejected')}»."
                    )
                )
                # уведомляем пользователя
                if user_obj:
                    await bot.send_message(
                        user_obj.id,
                        f"Ваше пополнение #{topup_id} на сумму {topup_obj.amount} руб. «{rus_status('rejected')}»."
                    )
        # обязательный ответ на callback_query
        return await bot.answer_callback_query(call.id)

    # ------------------------------------------------------------------------
    #            МОДЕРАЦИЯ ОТЗЫВОВ (approve/reject)
    # ------------------------------------------------------------------------
    @dp.callback_query(
        lambda call: call.data.startswith("approve_feedback_") or call.data.startswith("reject_feedback_")
    )
    async def handle_feedback_moderation(call: types.CallbackQuery):
        if not is_admin(call.from_user.id):
            return await bot.answer_callback_query(call.id, "Нет прав для модерации.", show_alert=True)

        feedback_id_str = call.data.split("_")[-1]
        try:
            feedback_id = int(feedback_id_str)
        except:
            return await bot.answer_callback_query(call.id, "Ошибка ID отзыва.", show_alert=True)

        with SessionLocal() as session:
            fb_obj = session.query(AdFeedback).filter_by(id=feedback_id).first()
            # Предположим, что у feedback есть поле status
            # Если нет — уберите проверку или адаптируйте
            if not fb_obj or getattr(fb_obj, "status", None) != "pending":
                return await bot.answer_callback_query(call.id, "Отзыв не найден или уже обработан.", show_alert=True)

            if call.data.startswith("approve_feedback_"):
                fb_obj.status = "approved"
                session.commit()
                await bot.answer_callback_query(call.id, "Отзыв одобрен.")
                return await bot.send_message(fb_obj.user_id, f"Ваш отзыв #{fb_obj.id} «{rus_status('approved')}»!")
            else:
                fb_obj.status = "rejected"
                session.commit()
                await bot.answer_callback_query(call.id, "Отзыв отклонён.")
                return await bot.send_message(fb_obj.user_id, f"Ваш отзыв #{fb_obj.id} «{rus_status('rejected')}».")

    # ------------------------------------------------------------------------
    #            ОДОБРИТЬ / ОТКЛОНИТЬ ВЫВОД СРЕДСТВ
    # ------------------------------------------------------------------------
    @dp.callback_query(
        lambda call: call.data.startswith("approve_withdraw_") or call.data.startswith("reject_withdraw_")
    )
    async def handle_withdraw_approval(call: types.CallbackQuery):
        if not is_admin(call.from_user.id):
            return await bot.answer_callback_query(call.id, "Нет прав для модерации.", show_alert=True)

        with SessionLocal() as session:
            if call.data.startswith("approve_withdraw_"):
                w_id_str = call.data.replace("approve_withdraw_", "")
                try:
                    w_id = int(w_id_str)
                except:
                    return await bot.answer_callback_query(call.id, "Некорректный ID вывода.", show_alert=True)

                wd = session.query(Withdrawal).filter_by(id=w_id, status="pending").first()
                if not wd:
                    return await bot.answer_callback_query(call.id, "Заявка не найдена или уже обработана.", show_alert=True)

                user_obj = session.query(User).filter_by(id=wd.user_id).first()
                if not user_obj:
                    return await bot.answer_callback_query(call.id, "Пользователь не найден.", show_alert=True)

                user_obj.balance = float(user_obj.balance) - float(wd.amount)
                wd.status = "approved"
                session.commit()

                await bot.answer_callback_query(call.id, "Вывод одобрен, баланс списан.")
                await bot.send_message(
                    call.message.chat.id,
                    f"✅Вывод #{w_id} «{rus_status('approved')}». С пользователя списано {wd.amount} руб."
                )
                return await bot.send_message(
                    wd.user_id,
                    f"Ваши средства ({wd.amount} руб.) отправлены на вывод!\n"
                    f"Баланс обновлён: {user_obj.balance} руб."
                )

            elif call.data.startswith("reject_withdraw_"):
                w_id_str = call.data.replace("reject_withdraw_", "")
                try:
                    w_id = int(w_id_str)
                except:
                    return await bot.answer_callback_query(call.id, "Некорректный ID вывода.", show_alert=True)

                wd = session.query(Withdrawal).filter_by(id=w_id, status="pending").first()
                if not wd:
                    return await bot.answer_callback_query(call.id, "Заявка не найдена или уже обработана.", show_alert=True)

                wd.status = "rejected"
                session.commit()

                await bot.answer_callback_query(call.id, "Вывод отклонён.")
                await bot.send_message(
                    call.message.chat.id,
                    f"❌Вывод #{w_id} «{rus_status('rejected')}»."
                )
                return await bot.send_message(
                    wd.user_id,
                    f"Ваша заявка на вывод #{w_id} «{rus_status('rejected')}»."
                )
            else:
                return None

    # ------------------------------------------------------------------------
    #            УПРАВЛЕНИЕ ПОДДЕРЖКОЙ (ТИКЕТАМИ)
    # ------------------------------------------------------------------------
    @dp.message(lambda m: m.text == "Управление поддержкой")
    async def admin_support_menu(message: types.Message):
        if not is_admin(message.chat.id):
            return
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, keyboard=[
            [
                types.KeyboardButton(text="Список открытых тикетов"),
                types.KeyboardButton(text="Главное меню")
            ]
        ])
        await bot.send_message(message.chat.id, "Управление тикетами поддержки:", reply_markup=kb)

    @dp.message(lambda m: m.text == "Список открытых тикетов")
    async def admin_list_tickets(message: types.Message):
        if not is_admin(message.chat.id):
            return None
        with SessionLocal() as session:
            tickets = session.query(SupportTicket).filter(SupportTicket.status == "open").all()
            if not tickets:
                return await bot.send_message(message.chat.id, "Нет открытых тикетов.")

            buttons = [
                [ types.InlineKeyboardButton(
                    text=f"Тикет #{t.id} от пользователя {t.user_id}",
                    callback_data=f"admin_support_view_{t.id}"
                ) ] for t in tickets
            ]
            kb = types.InlineKeyboardMarkup(inline_keyboard=buttons)
            return await bot.send_message(message.chat.id, "Открытые тикеты:", reply_markup=kb)

    # ------------------------------------------------------------------
    #   просмотр тикета (админ)
    # ------------------------------------------------------------------
    @dp.callback_query(lambda c: c.data.startswith("admin_support_view_"))
    async def admin_support_view_ticket(call: types.CallbackQuery):
        if not is_admin(call.from_user.id):
            return await bot.answer_callback_query(call.id, "Нет прав.")

        try:
            t_id = int(call.data.replace("admin_support_view_", "", 1))
        except ValueError:
            return await bot.answer_callback_query(call.id, "Некорректный ID тикета.", show_alert=True)

        with SessionLocal() as s:
            ticket = s.query(SupportTicket).get(t_id)
            if not ticket:
                return await bot.answer_callback_query(call.id, "Тикет не найден.", show_alert=True)

            text_history = "\n\n".join(
                f"{'Админ' if m.sender_id in ADMIN_IDS else f'Пользователь {m.sender_id}'} "
                f"({m.created_at:%d.%m.%y %H:%M}):\n{m.text}"
                for m in ticket.messages
            ) or "Сообщений пока нет."

        kb = types.InlineKeyboardMarkup(inline_keyboard=[
            [ types.InlineKeyboardButton(text="✉ Ответить", callback_data=f"admin_support_reply_{t_id}") ],
            [ types.InlineKeyboardButton(text="🛑 Закрыть тикет", callback_data=f"admin_support_close_{t_id}") ]
        ])
        await bot.edit_message_text(
            f"Тикет #{t_id}\nСтатус: {rus_status(ticket.status)}\n\n{text_history}",
            chat_id=call.message.chat.id, message_id=call.message.message_id, reply_markup=kb
        )
        return await bot.answer_callback_query(call.id)

    # ------------------------------------------------------------------
    #   ответ администратора
    # ------------------------------------------------------------------
    @dp.callback_query(lambda c: c.data.startswith("admin_support_reply_"))
    async def admin_support_reply_ticket(call: types.CallbackQuery, state: FSMContext):
        if not is_admin(call.from_user.id):
            return await bot.answer_callback_query(call.id)

        try:
            t_id = int(call.data.replace("admin_support_reply_", "", 1))
        except ValueError:
            return await bot.answer_callback_query(call.id, "Некорректный ID тикета.", show_alert=True)

        await bot.answer_callback_query(call.id)
        await state.set_state(AdminStates.reply_support_ticket)
        await state.update_data(t_id=t_id)
        return await bot.send_message(call.message.chat.id, f"Введите ответ для тикета #{t_id}:")

    @dp.message(AdminStates.reply_support_ticket)
    async def _admin_save_ticket_reply(message: types.Message, state: FSMContext):
        data = await state.get_data()
        await state.clear()
        t_id = data.get("t_id")
        """Сохраняет ответ админа и уведомляет пользователя."""
        text = (message.text or "").strip()
        if not text:
            return await bot.send_message(message.chat.id, "Пустое сообщение не отправлено.")

        # ── пишем в БД ─────────────────────────────────────────
        with SessionLocal() as s:
            tk = s.query(SupportTicket).get(t_id)
            if not tk or tk.status == "closed":
                return await bot.send_message(message.chat.id, "Тикет не найден или уже закрыт.")

            user_id = tk.user_id  # кешируем до commit/выхода
            s.add(SupportMessage(ticket_id=t_id,
                                 sender_id=message.chat.id,
                                 text=text))
            s.commit()

        # ── уведомляем пользователя ───────────────────────────
        try:
            await bot.send_message(
                user_id,
                f"[Поддержка] Администратор ответил в тикет #{t_id}:\n{text}"
            )
        except Exception:
            pass

        return await bot.send_message(message.chat.id, "Ответ отправлен.")

    # ------------------------------------------------------------------
    #   закрыть тикет (админ)
    # ------------------------------------------------------------------
    @dp.callback_query(lambda c: c.data.startswith("admin_support_close_"))
    async def admin_support_close_ticket(call: types.CallbackQuery):
        if not is_admin(call.from_user.id):
            return await bot.answer_callback_query(call.id, "Нет прав")

        try:
            t_id = int(call.data.replace("admin_support_close_", "", 1))
        except ValueError:
            return await bot.answer_callback_query(call.id, "Некорректный ID тикета.", show_alert=True)

        # сохраняем user_id до выхода из контекста
        with SessionLocal() as s:
            ticket = s.query(SupportTicket).get(t_id)
            if not ticket or ticket.status == "closed":
                return await bot.answer_callback_query(call.id, "Тикет не найден или уже закрыт.", show_alert=True)

            user_id = ticket.user_id  # ← кешируем!
            ticket.status = "closed"
            s.commit()

        await bot.answer_callback_query(call.id, "Тикет закрыт.")
        try:
            return await bot.send_message(user_id, f"Ваш тикет #{t_id} был закрыт администратором.")
        except Exception:
            return None

    # ------------------------------------------------------------------------
    #            ОБРАБОТКА ЖАЛОБ (complaint_msg_seller_, complaint_del_ad_, complaint_ban_)
    # ------------------------------------------------------------------------
    @dp.callback_query(lambda call:
        call.data.startswith("complaint_msg_seller_") or
        call.data.startswith("complaint_del_ad_") or
        call.data.startswith("complaint_ban_")
    )
    async def handle_complaint_actions(call: types.CallbackQuery, state: FSMContext):
        """
        Жалоба от search.py -> AdComplaint
        Кнопки:
          - «Написать продавцу»
          - «Удалить объявление»
          - «Заблокировать пользователя»
        """
        if not is_admin(call.from_user.id):
            return await bot.answer_callback_query(call.id, "Нет прав", show_alert=True)

        data = call.data
        if "complaint_msg_seller_" in data:
            c_id_str = data.replace("complaint_msg_seller_", "")
            action = "msg_seller"
        elif "complaint_del_ad_" in data:
            c_id_str = data.replace("complaint_del_ad_", "")
            action = "del_ad"
        else:
            c_id_str = data.replace("complaint_ban_", "")
            action = "ban_user"

        try:
            complaint_id = int(c_id_str)
        except:
            return await bot.answer_callback_query(call.id, "Некорректный ID жалобы.", show_alert=True)

        with SessionLocal() as session:
            comp = session.query(AdComplaint).filter_by(id=complaint_id).first()
            if not comp:
                return await bot.answer_callback_query(call.id, "Жалоба не найдена.", show_alert=True)

            ad_obj = session.query(Ad).filter_by(id=comp.ad_id).first()
            if not ad_obj:
                return await bot.answer_callback_query(call.id, "Объявление не найдено.", show_alert=True)

            comp.status = "in_progress"
            session.commit()

            seller_id = ad_obj.user_id

            if action == "msg_seller":
                # Написать продавцу
                await bot.answer_callback_query(call.id, "Введите сообщение продавцу (ответом на это).")
                await state.set_state(AdminStates.complaint_write_seller)
                await state.update_data(seller_id=seller_id)
                return await bot.send_message(
                    call.message.chat.id,
                    f"Напишите сообщение для продавца #{seller_id}:"
                )
            elif action == "del_ad":
                ad_obj.status = "rejected"
                comp.status = "resolved"
                session.commit()

                await bot.answer_callback_query(call.id, "Объявление отклонено/удалено.")
                return await bot.send_message(call.message.chat.id, f"Объявление #{ad_obj.id} -> 'rejected'.")
            elif action == "ban_user":
                await bot.answer_callback_query(call.id, "Укажите причину и срок (например: 'Мошенничество | 14').")
                await state.set_state(AdminStates.complaint_ban_user)
                await state.update_data(seller_id=seller_id, complaint_id=complaint_id)
                return await bot.send_message(
                    call.message.chat.id,
                    f"Введите причину и срок бана для пользователя #{seller_id}, формат:\n"
                    "`Причина | кол-во_дней` (пример: `Мошенничество | 7`)",
                    parse_mode="Markdown"
                )
            else:
                return None

    @dp.message(AdminStates.complaint_write_seller)
    async def process_msg_seller(message: types.Message, state: FSMContext):
        data = await state.get_data()
        await state.clear()
        seller_id = data.get("seller_id")
        text_to_seller = message.text.strip()
        try:
            await bot.send_message(seller_id, f"[Админ]: {text_to_seller}")
            await bot.send_message(message.chat.id, "Сообщение отправлено продавцу.")
        except:
            await bot.send_message(message.chat.id, "Ошибка при отправке (возможно, продавец не запустил бота).")

    @dp.message(AdminStates.complaint_ban_user)
    async def process_ban_user(message: types.Message, state: FSMContext):
        data = await state.get_data()
        await state.clear()
        seller_id = data.get("seller_id")
        complaint_id = data.get("complaint_id")
        txt = message.text.strip()
        if "|" not in txt:
            return await bot.send_message(message.chat.id, "Неверный формат. Нужно: `Причина | кол-во_дней`.")
        parts = txt.split("|", 1)
        reason = parts[0].strip()
        days_str = parts[1].strip()
        try:
            days_val = int(days_str)
        except:
            return await bot.send_message(message.chat.id, "Срок бана (в днях) не число.")

        with SessionLocal() as session:
            user_seller = session.query(User).filter_by(id=seller_id).first()
            if not user_seller:
                return await bot.send_message(message.chat.id, "Продавец не найден.")

            user_seller.is_banned = True
            user_seller.ban_reason = reason
            dt_until = datetime.now(timezone.utc) + timedelta(days=days_val)
            user_seller.ban_until = dt_until

            comp = session.query(AdComplaint).filter_by(id=complaint_id).first()
            if comp:
                comp.status = "resolved"

            session.commit()

        return await bot.send_message(
            message.chat.id,
            f"Пользователь {seller_id} заблокирован.\nПричина: {reason}.\nСрок: {days_val} дн."
        )

    # ============================
    #    Редактировать профиль пользователя
    # ============================
    @dp.message(lambda m: m.text == "Редактировать профиль пользователя")
    async def edit_profile_user_start(message: types.Message, state: FSMContext):
        if not is_admin(message.chat.id):
            return
        await state.set_state(AdminStates.edit_profile_user_id)
        await bot.send_message(
            message.chat.id,
            "Введите ID пользователя, профиль которого хотите изменить:"
        )

    @dp.message(AdminStates.edit_profile_user_id)
    async def process_edit_profile_user_id(message: types.Message, state: FSMContext):
        chat_id = message.chat.id
        try:
            uid = int(message.text.strip())
        except:
            await state.clear()
            return await bot.send_message(chat_id, "Некорректный ID.")

        with SessionLocal() as session:
            user = session.query(User).filter_by(id=uid).first()
            if not user:
                await state.clear()
                return await bot.send_message(chat_id, f"Пользователь #{uid} не найден.")

        await state.update_data(uid=uid)
        await state.set_state(AdminStates.edit_profile_field)
        return await bot.send_message(
            chat_id,
            f"Редактирование профиля пользователя #{uid}.\n"
            "Введите одно из: *fio* (ФИО), *inn* (ИНН), *company* (название компании).",
            parse_mode="Markdown"
        )

    @dp.message(AdminStates.edit_profile_field)
    async def process_edit_profile_field(message: types.Message, state: FSMContext):
        data = await state.get_data()
        chat_id = message.chat.id
        field = message.text.strip().lower()
        if field not in ["fio", "inn", "company"]:
            await state.clear()
            return await bot.send_message(chat_id, "Неизвестное поле. Операция прервана.")

        await state.update_data(field=field)
        await state.set_state(AdminStates.edit_profile_value)
        return await bot.send_message(chat_id, f"Введите новое значение поля *{field.upper()}*:", parse_mode="Markdown")

    @dp.message(AdminStates.edit_profile_value)
    async def process_edit_profile_value(message: types.Message, state: FSMContext):
        data = await state.get_data()
        await state.clear()
        chat_id = message.chat.id
        new_val = message.text.strip()
        user_id = data.get("uid")
        field = data.get("field")
        with SessionLocal() as session:
            user = session.query(User).filter_by(id=user_id).first()
            if not user:
                return await bot.send_message(chat_id, f"Пользователь #{user_id} не найден при обновлении.")

            if field == "fio":
                user.full_name = new_val
            elif field == "inn":
                user.inn = new_val
            elif field == "company":
                user.company_name = new_val
            else:
                return await bot.send_message(chat_id, "Неизвестное поле. Прервано.")

            session.commit()

        return await bot.send_message(chat_id, f"Поле {field.upper()} пользователя #{user_id} обновлено на: {new_val}")

    @dp.callback_query(lambda call:
    call.data.startswith("approve_ext_") or call.data.startswith("reject_ext_")             )
    async def handle_extension_request(call: types.CallbackQuery):
        """
        Админ одобряет или отклоняет просьбу о продлении.
        """
        admin_id = call.from_user.id
        if not is_admin(admin_id):
            return await bot.answer_callback_query(call.id, "Нет прав.", show_alert=True)

        data, ad_id_str = call.data.split("_", 1)[0:2], call.data.split("_", 2)[2]
        ad_id = int(ad_id_str)

        with SessionLocal() as sess:
            ad = sess.query(Ad).get(ad_id)
            if not ad:
                return await bot.answer_callback_query(call.id, "Объявление не найдено.", show_alert=True)

            if call.data.startswith("approve_ext_"):
                # сдвигаем created_at на сейчас
                ad.created_at = datetime.now(timezone.utc)
                sess.commit()

                # уведомляем
                await bot.edit_message_reply_markup(
                    chat_id=call.message.chat.id, message_id=call.message.message_id, reply_markup=None
                )
                await bot.send_message(admin_id, f"✅ Продление объявления #{ad_id} одобрено.")
                await bot.send_message(
                    ad.user_id,
                    f"Ваше объявление #{ad_id} продлено на 30 дней!"
                )
            else:
                # отклоняем
                await bot.edit_message_reply_markup(
                    chat_id=call.message.chat.id, message_id=call.message.message_id, reply_markup=None
                )
                await bot.send_message(admin_id, f"❌ Продление объявления #{ad_id} отклонено.")
                await bot.send_message(
                    ad.user_id,
                    f"К сожалению, продление объявления #{ad_id} отклонено администратором."
                )

        return await bot.answer_callback_query(call.id)
